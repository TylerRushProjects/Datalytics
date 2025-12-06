import os
import pandas as pd
from core.state import get_dataframe, get_duplicate_highlight, get_data_version
from utils.menus import show_export_menu
from core.audit import log_action, save_audit_log_to_txt


def export_flow():
    """
    Top-level export flow:
    1) Check there is an active DataFrame.
    2) Ask export type (CSV or XLSX).
    3) Prompt for file path.
    4) Perform export.
    """
    df = get_dataframe()

    if df is None:
        print("No data to export. Please import and transform a file first.")
        return

    while True:
        choice = show_export_menu()

        if choice == "0":
            return

        if choice in ("1", "2"):
            path = input("Enter full export file path (including filename and extension): ").strip()
            if not path:
                print("No path provided. Export cancelled.")
                return

            if choice == "1":
                _export_csv(df, path)
            elif choice == "2":
                _export_xlsx(df, path)

            # After a successful export attempt, return to main menu
            return

        print("Invalid export option. Please try again.")


def _export_csv(df: pd.DataFrame, path: str) -> None:
    """Export DataFrame as CSV."""
    if not path.lower().endswith(".csv"):
        print("Warning: Path does not end with .csv; appending extension.")
        path += ".csv"

    try:
        directory = os.path.dirname(path)
        if directory:
            os.makedirs(directory, exist_ok=True)
        df.to_csv(path, index=False)
        print(f"CSV export complete: {path}")
    except Exception as e:
        print(f"Error during CSV export: {e}")

        log_action(
            "EXPORT_CSV",
            details=f"Exported CSV to '{path}'.",
            rows_affected=len(df),
        )

        maybe_save_audit_log(path)


def _export_xlsx(df: pd.DataFrame, path: str) -> None:
    """
    Export DataFrame as XLSX.
    If duplicate highlight configuration exists (from DUP-1), apply highlighting.
    """
    if not path.lower().endswith(".xlsx"):
        print("Warning: Path does not end with .xlsx; appending extension.")
        path += ".xlsx"

    try:
        directory = os.path.dirname(path)
        if directory:
            os.makedirs(directory, exist_ok=True)

        # Base export
        df.to_excel(path, index=False, engine="openpyxl")
        print(f"Base XLSX export complete: {path}")

        # Try to apply highlighting based on prior DUP-1 config
        _apply_duplicate_highlighting(path, df)

        log_action(
            "EXPORT_XLSX",
            details=f"Exported XLSX to '{path}'.",
            rows_affected=len(df),
        )

        maybe_save_audit_log(path)

    except Exception as e:
        print(f"Error during XLSX export: {e}")


def _apply_duplicate_highlighting(path: str, df: pd.DataFrame) -> None:
    """
    If duplicate highlight info is available from DUP-1 and still matches
    the current data_version, apply cell highlighting. Otherwise skip.
    """
    highlight_info = get_duplicate_highlight()
    if not highlight_info:
        print("No duplicate highlight configuration found. XLSX exported without highlighting.")
        return

    # Check if the data has changed since Identify
    current_version = get_data_version()
    saved_version = highlight_info.get("data_version")

    if saved_version is None or saved_version != current_version:
        print("Duplicate highlight configuration is stale. XLSX exported without highlighting.")
        return

    columns = highlight_info.get("columns", [])
    dup_indices = highlight_info.get("duplicate_index", [])

    if not columns or not dup_indices:
        print("Duplicate highlight info incomplete. XLSX exported without highlighting.")
        return

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb = load_workbook(path)
        ws = wb.active

        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        col_to_excel_index = {
            col_name: idx + 1 for idx, col_name in enumerate(df.columns)
        }

        for idx in dup_indices:
            try:
                pos = df.index.get_loc(idx)  # 0-based row position in current DataFrame
            except KeyError:
                continue

            excel_row = pos + 2  # +1 header row, +1 for 1-based index

            for col_name in columns:
                excel_col = col_to_excel_index.get(col_name)
                if excel_col is None:
                    continue
                cell = ws.cell(row=excel_row, column=excel_col)
                cell.fill = fill

        wb.save(path)
        print("Duplicate highlighting applied to XLSX export.")

    except ImportError:
        print("openpyxl is required for highlighting but not installed. XLSX exported without highlighting.")
    except Exception as e:
        print(f"Error applying duplicate highlighting: {e}")


def maybe_save_audit_log(export_path: str):
    print("\nWould you like to save the audit log as a text file in the same folder?")
    print("1. Yes")
    print("2. No")
    choice = input("Enter choice: ").strip()

    if choice == "1":
        save_audit_log_to_txt(export_path)
    else:
        print("Audit log not saved.")